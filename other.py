import os
import shutil
import threading
from datetime import datetime
from time import sleep

import pandas as pd
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from Outlook import Outlook
from database.database import DataBase


def get_date_and_time() -> str:
    """
    Возвращает текущую дату и время

    Returns:
        str: 24.06.2024__16-50-33
    """
    return datetime.now().strftime("%d.%m.%Y__%H-%M-%S")


def get_full_path_to_folder(date_and_time: str = None, email: str = None, new_folder: str = None) -> str:
    """
    Проверяет на существование папки, если такой нет, то создаёт ее

    Args:
        date_and_time (str): дата и время
        email (str): почта, на которую нужно будет отправить данные
        new_folder (str): полный путь до новой директории

    Returns:
        str: .../User_Data/r.purdyshev@lcgs.ru/24.06.2024__16:50:33
    """
    if email:
        path_for_save_excel = os.path.join(os.getcwd(), 'user_data', email, date_and_time)
        if not os.path.exists(path_for_save_excel):
            os.makedirs(path_for_save_excel, exist_ok=True)
        return path_for_save_excel
    else:
        path_to_folder = os.path.join(new_folder)
        if not os.path.exists(path_to_folder):
            os.makedirs(path_to_folder, exist_ok=True)
        return path_to_folder


def write_data_in_database(email: str, date_and_time: str, path_to_file: str) -> None:
    """
    Запись данных в БД

    :param email: email пользователя, который запросил выгрузку
    :param date_and_time: дата и время
    :param path_to_file: путь до Excel файла
    """
    db = DataBase()
    db.add_user(
        user_email=email,
        date_and_time=date_and_time,
        path_to_file=path_to_file,
        status='start'
    )
    db.close_connection()


def save_user_file(filename: str, email: str, file: FileStorage) -> str:
    """
    Сохранение Excel, который прикрепил пользователь на Web интерфейсе

    :param filename: имя файла
    :param email: email пользователя
    :param file: объект файла (сам файл)

    :return: path_to_file - полный путь до файла
    """
    date_and_time = get_date_and_time()

    path_to_folder_user = get_full_path_to_folder(email=email, date_and_time=date_and_time)
    path_to_file = os.path.join(path_to_folder_user, filename)
    file.save(path_to_file)

    write_data_in_database(
        email=email,
        date_and_time=date_and_time,
        path_to_file=path_to_folder_user
    )

    return path_to_file


def update_last_column_excel(
        path_to_excel: str,
        full_name: str,
        inn_on_site: str,
        passport_data: str,
        log: object = None,
        lock: threading.Lock = None
):
    """
    Изменяет значение столбца "Проверка" и "Новый ИНН" в Excel

    :param path_to_excel: путь до Excel файла
    :param full_name: ФИО клиента
    :param inn_on_site: новый ИНН с сайта
    :param passport_data: паспортные данные клиента
    :param log: объект класса логирования
    :param lock: объект блокировки для синхронизации доступа к файлу
    """
    if lock:
        with lock:
            if log:
                log.info("Начинаю менять значения в файле")

            try:
                wb = load_workbook(path_to_excel)
                ws = wb.active

                # Проверка наличия столбца "ИНН на сайте"
                if 'ИНН на сайте' not in [cell.value for cell in ws[1]]:
                    # Добавление столбца
                    ws.cell(row=1, column=ws.max_column + 1, value='ИНН на сайте')
                    wb.save(path_to_excel)

                # Поиск строки с нужным значением
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    new_row = [str(_cell) for _cell in row]

                    if full_name in new_row:
                        passport_data_excel = ''.join(new_row[-3:-1])
                        if passport_data == passport_data_excel:
                            check_column_inn = [cell.value for cell in ws[1]].index('ИНН на сайте') + 1
                            ws.cell(row=row_num, column=check_column_inn, value=inn_on_site)

                            wb.save(path_to_excel)

                            if log:
                                log.info("Файл успешно обновлен!")
                            break
                        else:
                            continue
                else:
                    if log:
                        log.info(f"Строка с ФИО '{full_name}' не найдена. Не могу изменить значение ячейки")

                sleep(3)
            except Exception as e:
                if log:
                    log.error(f"Ошибка при обновлении файла: {e}")
    else:
        if log:
            log.error("Объект блокировки не передан. Не могу синхронизировать доступ к файлу.")


def checking_correctness_excel(dataframe: list, email: str) -> bool:
    check_list = ['ФИО', 'Дата рождения', 'Серия паспорта', 'Номер паспорта']

    if check_list == dataframe[-4:]:
        return True
    else:
        Outlook().send_mail_error(mail_to=[email])
        exit()


def get_data_from_excel(path_to_excel: str, email: str) -> list[list[str]]:
    """
    Получение данных из Excel файла

    :param path_to_excel: полный путь до Excel файла
    :return: список в списке, где внутренний список является строкой из Excel файла
    """
    try:
        df = pd.read_excel(path_to_excel, dtype=str)

        if checking_correctness_excel(dataframe=df.keys().tolist(), email=email):
            filter_columns = ((df['ФИО'] != '') & (df['ФИО'] != 'nan') &
                              (df['Дата рождения'] != '') & (df['Дата рождения'] != 'nan') &
                              (df['Серия паспорта'] != '') & (df['Серия паспорта'] != 'nan') &
                              (df['Номер паспорта'] != '') & (df['Номер паспорта'] != 'nan'))

            filtered_df = df[filter_columns]

            return filtered_df.values.tolist()

    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")

        return []


def check_null_clients(path_to_file: str):
    """
    Проверка пустых значений в столбце "ИНН на сайте"

    :param path_to_file: путь до Excel файла

    :return: список клиентов, которые не прошли обработку в первый раз
    """

    df = pd.read_excel(path_to_file, dtype=str)
    filter_columns = (df['ИНН на сайте'] == '') | (df['ИНН на сайте'].isna())
    filtered_df = df[filter_columns]
    result = [row[0:-1] for row in filtered_df.values.tolist()]

    return result


def copy_result_folder_to_public_folder(path_to_local_folder: str, email: str, log: object = None) -> str | None:
    """
    Копирование итоговых файлов в публичную директорию

    :param path_to_local_folder: путь до директории, которую нужно перенести
    :param log: объект класса логгирования (по умолчанию None)
    """
    logger = print
    if log:
        logger = log.info

    logger('Начинаю копировать готовые файлы в публичную директорию')
    folder_name = os.path.basename(path_to_local_folder)

    path_to_public_folder = r'S:\В отдел\Отдел судебного взыскания\Проверка ИНН'
    full_path_to_user_folder = os.path.join(path_to_public_folder, email, folder_name)

    try:
        shutil.copytree(path_to_local_folder, full_path_to_user_folder)
        logger('Копирование прошло успешно')
        return full_path_to_user_folder
    except Exception as ex:
        logger(f'Произошла ошибка при копировании файлов на общий диск! Ошибка:\n{ex}')
        return None
